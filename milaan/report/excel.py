"""Excel exception report — the deliverable an accountant actually opens.

Sheets:
  Summary          — counts, match rate, what needs attention
  Matched          — full audit trail (rule, score, explanation)
  Unmatched Bank   — bank lines needing review, with near-miss candidates
  Unmatched Ledger — ledger entries with no bank movement
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

HDR = Font(bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="0B5D3B")
WARN = PatternFill("solid", fgColor="FFF3CD")


def _sheet(wb, title, headers, widths):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c, w in enumerate(widths, 1):
        ws.cell(1, c).font = HDR
        ws.cell(1, c).fill = FILL
        ws.column_dimensions[ws.cell(1, c).column_letter].width = w
    ws.freeze_panes = "A2"
    return ws


def write_report(path, result, txns, entries, meta=None, categories=None):
    tmap = {t.txn_id: t for t in txns}
    emap = {e.entry_id: e for e in entries}
    wb = Workbook()
    wb.remove(wb.active)

    # ---------------- Summary
    ws = _sheet(wb, "Summary", ["Item", "Value"], [42, 24])
    n_bank, n_ledger = len(txns), len(entries)
    n_matched_bank = sum(len(m["bank_ids"]) for m in result["matches"])
    n_matched_ledger = sum(len(m["ledger_ids"]) for m in result["matches"])
    dupes = [u for u in result["unmatched_bank"] if u.get("suspect") == "DUPLICATE"]
    rows = [
        ("Bank transactions", n_bank),
        ("Ledger entries", n_ledger),
        ("Bank txns matched", f"{n_matched_bank} ({n_matched_bank/n_bank:.1%})"),
        ("Ledger entries matched", f"{n_matched_ledger} ({n_matched_ledger/max(n_ledger,1):.1%})"),
        ("Suspected duplicate bank lines", len(dupes)),
        ("Bank lines needing review", len(result["unmatched_bank"]) - len(dupes)),
        ("Ledger entries with no bank movement", len(result["unmatched_ledger"])),
    ]
    if meta:
        rows = list(meta.items()) + rows
    for r in rows:
        ws.append(list(r))

    # ---------------- Matched (audit trail)
    ws = _sheet(wb, "Matched",
                ["Rule", "Bank txn(s)", "Bank date", "Bank amount",
                 "Ledger entry(ies)", "Party", "Ledger amount", "Explanation"],
                [14, 16, 12, 14, 18, 22, 14, 60])
    for m in sorted(result["matches"], key=lambda x: x["bank_ids"][0]):
        bts = [tmap[b] for b in m["bank_ids"]]
        les = [emap[l] for l in m["ledger_ids"]]
        ws.append([m["rule"],
                   ", ".join(m["bank_ids"]), bts[0].date,
                   sum(t.amount for t in bts),
                   ", ".join(m["ledger_ids"]), les[0].party,
                   sum(e.amount for e in les), m["explanation"]])
        if m["rule"] != "P1_EXACT":
            for c in range(1, 9):
                ws.cell(ws.max_row, c).fill = WARN

    # ---------------- Unmatched Bank
    ws = _sheet(wb, "Unmatched Bank",
                ["Bank txn", "Date", "Direction", "Amount", "Narration",
                 "Category", "Suspect",
                 "Nearest candidates (entry / party / delta / name score)"],
                [10, 12, 10, 14, 42, 16, 22, 60])
    for u in result["unmatched_bank"]:
        t = tmap[u["bank_id"]]
        if u.get("suspect") == "DUPLICATE":
            suspect = f"DUPLICATE of {u['duplicate_of']}"
            near = ""
        else:
            suspect = ""
            near = " | ".join(
                f"{c['entry_id']} {c['party']} Δ{c['delta']:,.2f} ns{c['name_score']:.0f}"
                for c in u.get("near_misses", []))
        cat = (categories or {}).get(t.txn_id, "")
        ws.append([t.txn_id, t.date, t.direction, t.amount,
                   t.narration, cat, suspect, near])

    # ---------------- Unmatched Ledger
    ws = _sheet(wb, "Unmatched Ledger",
                ["Entry", "Date", "Party", "Description", "Invoice",
                 "Amount", "Type"],
                [10, 12, 22, 30, 12, 14, 10])
    for u in result["unmatched_ledger"]:
        e = emap[u["ledger_id"]]
        ws.append([e.entry_id, e.date, e.party, e.description,
                   e.invoice_no, e.amount, e.entry_type])

    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=(cell.column in (5, 7, 8)))
    wb.save(path)
    return path
