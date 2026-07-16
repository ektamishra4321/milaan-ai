"""Week 2 tests — matching engine, eval harness, Excel report. No keys."""

import json

from milaan.generator.generate import generate_month, write_month
from milaan.ingest.adapters import parse_statement, parse_ledger
from milaan.match.engine import (reconcile, name_score, narration_hint,
                                 flag_duplicates)
from milaan.evalx.harness import evaluate_month
from milaan.report.excel import write_report


def _run(seed=101, fmt="hdfc", tmp_path=None):
    bank, ledger, gt = generate_month(seed=seed, year=2026, month=3)
    if tmp_path:
        write_month(str(tmp_path), fmt, bank, ledger, gt)
        txns, _, _ = parse_statement(str(tmp_path / f"statement_{fmt}.csv"))
        entries = parse_ledger(str(tmp_path / "ledger.csv"))
        return txns, entries, gt
    return bank, ledger, gt


# ------------------------------------------------------------ narration/name
def test_narration_hint_strips_noise():
    h = narration_hint("UPI-SHRM TRDRS-sharma4@okicici-428730877482-PAYMENT")
    assert h == "SHRM TRDRS"


def test_name_score_handles_vowel_drop():
    assert name_score("IMPS-371052246567-SHRMA TRDRS", "Sharma Traders") >= 85


def test_name_score_rejects_wrong_party():
    assert name_score("NEFT CR-N12345678-BALAJI PACKAGING", "Sharma Traders") < 70


# ------------------------------------------------------------ duplicates
def test_flag_duplicates():
    bank, _, gt = _run(seed=205)
    dupes = flag_duplicates(bank)
    truth = {u["bank_id"] for u in gt["unmatched_bank"]
             if u["reason"] == "DUPLICATE_BANK"}
    assert truth <= (set(dupes.keys()) | set())  # every planted dupe flagged


# ------------------------------------------------------------ engine quality
def test_engine_meets_bar_on_fresh_seeds(tmp_path):
    """Seeds never used during tuning; PRD bar: P>=0.98, R>=0.90."""
    agg = {"pred": 0, "truth": 0, "tp": 0}
    for i, seed in enumerate([911, 912, 913]):
        d = tmp_path / str(seed)
        d.mkdir()
        txns, entries, gt = _run(seed=seed, fmt=["hdfc", "icici", "sbi"][i],
                                 tmp_path=d)
        res = reconcile(txns, entries)
        m = evaluate_month(res, gt)
        agg["pred"] += m["pairs_pred"]
        agg["truth"] += m["pairs_truth"]
        agg["tp"] += m["pairs_correct"]
    precision = agg["tp"] / agg["pred"]
    recall = agg["tp"] / agg["truth"]
    assert precision >= 0.98, f"precision {precision:.3f} below bar"
    assert recall >= 0.90, f"recall {recall:.3f} below bar"


def test_engine_deterministic():
    bank, ledger, _ = _run(seed=77)
    a = reconcile(bank, ledger)
    b = reconcile(bank, ledger)
    assert json.dumps(a, sort_keys=True) == json.dumps(b, sort_keys=True)


def test_every_match_has_audit_trail():
    bank, ledger, _ = _run(seed=88)
    res = reconcile(bank, ledger)
    for m in res["matches"]:
        assert m["rule"] in {"P1_EXACT", "P2_TOLERANCE", "P3_COMBINED", "P4_SPLIT"}
        assert m["explanation"]


def test_no_double_assignment():
    bank, ledger, _ = _run(seed=99)
    res = reconcile(bank, ledger)
    bank_ids, ledger_ids = [], []
    for m in res["matches"]:
        bank_ids += m["bank_ids"]
        ledger_ids += m["ledger_ids"]
    bank_ids += [u["bank_id"] for u in res["unmatched_bank"]]
    ledger_ids += [u["ledger_id"] for u in res["unmatched_ledger"]]
    assert sorted(bank_ids) == sorted(t.txn_id for t in bank)
    assert sorted(ledger_ids) == sorted(e.entry_id for e in ledger)


# ------------------------------------------------------------ report
def test_excel_report_writes(tmp_path):
    txns, entries, _ = _run(seed=55, fmt="sbi", tmp_path=tmp_path)
    res = reconcile(txns, entries)
    out = tmp_path / "report.xlsx"
    write_report(str(out), res, txns, entries, meta={"Client": "Test & Co"})
    assert out.exists() and out.stat().st_size > 5000
    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert set(wb.sheetnames) == {"Summary", "Matched", "Unmatched Bank",
                                  "Unmatched Ledger"}


def test_report_with_categories(tmp_path):
    txns, entries, _ = _run(seed=66, fmt="hdfc", tmp_path=tmp_path)
    res = reconcile(txns, entries)
    cats = {u["bank_id"]: "bank_charges" for u in res["unmatched_bank"]}
    out = tmp_path / "r.xlsx"
    write_report(str(out), res, txns, entries, categories=cats)
    from openpyxl import load_workbook
    ws = load_workbook(out)["Unmatched Bank"]
    headers = [c.value for c in ws[1]]
    assert "Category" in headers
